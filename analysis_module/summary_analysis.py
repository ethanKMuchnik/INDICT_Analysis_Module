import pandas as pd
from openpyxl.styles import Font, PatternFill

def conglomerate_patient_data(output_dict, reference_data_key,metric_column ='daily_SD_rate'):
	"""
	Conglomerates bucketed data across patients into treatment and standard group tables.

	Args:
		output_dict: The output dictionary from INDICT_XLSX_Analysis
		metric_column: The column from bucketed_events_df to use
					  (e.g., 'daily_SD_rate', 'tier_character', 'num_events', 'valid_recording_hours')

	Returns:
		treatment_table: DataFrame with time_hours as first column, then metric for each treatment patient
		standard_table: DataFrame with time_hours as first column, then metric for each standard patient
	"""

	treatment_patients = {}
	standard_patients = {}

	# Separate patients by treatment group
	for patient_id, patient_data in output_dict.items():
		bucketed_df = patient_data[reference_data_key][['time_hours', metric_column]].copy()
		bucketed_df = bucketed_df.rename(columns={metric_column: patient_id})

		if patient_data['patient_treatment_group'] == 'Treatment':
			treatment_patients[patient_id] = bucketed_df
		else:
			standard_patients[patient_id] = bucketed_df

	# Build treatment table
	if treatment_patients:
		treatment_table = None
		for patient_id, df in treatment_patients.items():
			if treatment_table is None:
				treatment_table = df
			else:
				treatment_table = treatment_table.merge(df, on='time_hours', how='outer')
	else:
		treatment_table = pd.DataFrame()

	# Build standard table
	if standard_patients:
		standard_table = None
		for patient_id, df in standard_patients.items():
			if standard_table is None:
				standard_table = df
			else:
				standard_table = standard_table.merge(df, on='time_hours', how='outer')
	else:
		standard_table = pd.DataFrame()

	return treatment_table, standard_table


def export_conglomerated_data(output_dict,reference_data_key, save_path ,metric_columns=['daily_SD_rate', 'tier_character', 'num_events', 'valid_recording_hours']):
	"""
	Exports conglomerated patient data to an Excel file with separate sheets for each metric.

	Args:
		output_dict: The output dictionary from INDICT_XLSX_Analysis
		save_path: Path where to save the Excel file (should end with .xlsx)
		metric_columns: List of metric columns to export (default: all common metrics)
	"""

	# Create Excel writer
	excel_writer = pd.ExcelWriter(save_path, engine='openpyxl')

	# Process each metric
	for metric in metric_columns:
		treatment_table, standard_table = conglomerate_patient_data(output_dict, reference_data_key,metric_column=metric)

		# Write treatment table
		if not treatment_table.empty:
			treatment_table.to_excel(excel_writer, sheet_name=f'{metric}_treatment', index=False)

			# Format the sheet
			worksheet = excel_writer.sheets[f'{metric}_treatment']

			# Bold header row
			for cell in worksheet[1]:
				cell.font = Font(bold=True)
				cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

		# Write standard table
		if not standard_table.empty:
			standard_table.to_excel(excel_writer, sheet_name=f'{metric}_standard', index=False)

			# Format the sheet
			worksheet = excel_writer.sheets[f'{metric}_standard']

			# Bold header row
			for cell in worksheet[1]:
				cell.font = Font(bold=True)
				cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

	# Close the Excel writer
	excel_writer.close()

	print(f"Conglomerated data saved to: {save_path}")


def export_tier_summary_data(output_dict, save_path):
	"""
	Exports tier-conglomerated data across all patients to an Excel file.
	Creates a sheet with columns for each patient and rows for tier metrics.

	Args:
		output_dict: The output dictionary from INDICT_XLSX_Analysis
		save_path: Path where to save the Excel file (should end with .xlsx)
	"""

	# Define the rows we want to extract
	row_definitions = [
		('Total Events', 'All', 'num_events'),
		('Tier 1 Events', 'Tier1', 'num_events'),
		('Tier 2 Events', 'Tier2', 'num_events'),
		('Tier 3 Events', 'Tier3', 'num_events'),
		('Total Valid Time (hrs)', 'All', 'valid_hours'),
		('Tier 1 Valid Time (hrs)', 'Tier1', 'valid_hours'),
		('Tier 2 Valid Time (hrs)', 'Tier2', 'valid_hours'),
		('Tier 3 Valid Time (hrs)', 'Tier3', 'valid_hours'),
		('Total Daily Rate', 'All', 'daily_SD_rate'),
		('Tier 1 Daily Rate', 'Tier1', 'daily_SD_rate'),
		('Tier 2 Daily Rate', 'Tier2', 'daily_SD_rate'),
		('Tier 3 Daily Rate', 'Tier3', 'daily_SD_rate'),
	]

	# Get all patient IDs
	patient_ids = list(output_dict.keys())

	# Create data dictionary with row labels as index
	data = {'Metric': [row_def[0] for row_def in row_definitions]}

	# Extract data for each patient
	for patient_id in patient_ids:
		patient_data = output_dict[patient_id]
		patient_column = []

		for row_label, tier_key, metric_key in row_definitions:
			# Access the data from Summary dict
			if tier_key in patient_data['Summary']:
				value = patient_data['Summary'][tier_key].get(metric_key, None)
				patient_column.append(value)
			else:
				patient_column.append(None)

		data[patient_id] = patient_column

	# Create DataFrame
	df = pd.DataFrame(data)

	# Create Excel writer
	excel_writer = pd.ExcelWriter(save_path, engine='openpyxl')

	# Write to sheet
	df.to_excel(excel_writer, sheet_name='Tier Summary', index=False)

	# Format the sheet
	worksheet = excel_writer.sheets['Tier Summary']

	# Bold header row
	for cell in worksheet[1]:
		cell.font = Font(bold=True)
		cell.fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')

	# Bold first column (Metric names)
	for row in worksheet.iter_rows(min_row=2, max_row=len(row_definitions)+1, min_col=1, max_col=1):
		for cell in row:
			cell.font = Font(bold=True)

	# Close the Excel writer
	excel_writer.close()

	print(f"Tier summary data saved to: {save_path}")
